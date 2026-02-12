"""
Create a notation/parameter table for the Implicit-Explicit Diffusion Model paper.
Style: Matching TABLE I (Notations and Explanations) from published paper reference.
"""
import openpyxl
from openpyxl.styles import (Font, Alignment, Border, Side, PatternFill,
                               NamedStyle)
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Notations and Explanations"

# ── Style definitions ──
# Fonts
ft_title    = Font(name='Times New Roman', size=14, bold=True, color='000000')
ft_subtitle = Font(name='Times New Roman', size=11, bold=True, color='000000')
ft_header   = Font(name='Times New Roman', size=11, bold=True, color='000000')
ft_body     = Font(name='Times New Roman', size=11, italic=True, color='000000')
ft_body_r   = Font(name='Times New Roman', size=11, italic=False, color='000000')

# Alignment
al_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
al_left   = Alignment(horizontal='left', vertical='center', wrap_text=True)

# Borders  (mimic the thick top/bottom rule + thin header separator style)
thick = Side(style='thick', color='000000')
medium = Side(style='medium', color='000000')
thin  = Side(style='thin', color='000000')
none  = Side(style=None)

bdr_top    = Border(top=thick)
bdr_bottom = Border(bottom=thick)
bdr_top_bottom_thick = Border(top=thick, bottom=thick)
bdr_header = Border(bottom=medium)
bdr_none   = Border()

# Header fill (soft red/salmon to match reference "TABLE I" highlight bar)
fill_header = PatternFill(start_color='E8C4C4', end_color='E8C4C4', fill_type='solid')

# Column widths
ws.column_dimensions['A'].width = 3    # left margin
ws.column_dimensions['B'].width = 22   # Notation column
ws.column_dimensions['C'].width = 55   # Explanation column
ws.column_dimensions['D'].width = 3    # right margin

# ── Data: Notation → Explanation ──
notations = [
    # General symbols
    ("Z",               "Original time series data"),
    ("Z_θ",             "Time series output from model inference"),
    ("Z̃",              "Missing / corrupted time series data"),
    ("M",               "Binary observation mask (1 = observed, 0 = missing)"),
    ("x",               "Noisy input to the denoising network"),
    ("ε_θ",             "Noise predicted by the model"),
    # Architecture dimensions
    ("D",               "Input feature dimension (number of variates)"),
    ("H",               "Hidden representation dimension (default 128)"),
    ("L",               "Length of time series (number of timesteps)"),
    ("K",               "Number of sensor nodes / traffic channels"),
    ("N",               "State dimension in S4 layer (default 256)"),
    ("C",               "Number of feature channels"),
    # Diffusion parameters
    ("T",               "Total number of diffusion steps"),
    ("β_t",             "Noise schedule at diffusion step t"),
    ("t",               "Current diffusion timestep index"),
    ("T_emb",           "Sinusoidal embedding for diffusion timestep"),
    # Implicit module (Dilated Causal Conv)
    ("F_imp",           "Multiscale implicit feature set"),
    ("d_i",             "Dilation factor of the i-th conv layer (2^i)"),
    ("k",               "Convolution kernel size (default 5)"),
    ("P",               "Causal padding size: (k − 1) × d"),
    ("α_i",             "Learnable residual weight for conv layer i"),
    ("h_i",             "Hidden features after the i-th conv layer"),
    # Explicit module (S4 Layer)
    ("F_exp",           "Explicit feature set from structured state space"),
    ("A",               "State transition matrix (HiPPO-initialized)"),
    ("B",               "Input-to-state projection matrix"),
    ("C",               "State-to-output readout matrix"),
    ("D",               "Direct feedthrough (skip) vector"),
    ("Ā, B̄",           "Discretized state-space matrices (bilinear)"),
    ("s_t",             "Latent state vector at timestep t"),
    ("Δ",               "Discretization timestep (learnable log-scale)"),
    ("n_heads",         "Number of parallel S4 heads (default 2)"),
    ("d_head",          "Per-head state dimension: N / n_heads"),
    # Gated Fusion
    ("σ",               "Sigmoid activation for gating"),
    ("g",               "Gating coefficient: σ(W·[F_imp; F_exp] + b)"),
    # Mask Embedding
    ("E_mask",          "Learned binary embedding for mask tokens"),
    ("r",               "Missing ratio per timestep (mean of mask)"),
    ("PE",              "Sinusoidal positional encoding"),
    # Residual blocks
    ("R",               "Number of residual blocks (default 6)"),
    ("p",               "Dropout probability (default 0.15)"),
]

# ══════════════════════════════════════════════════════════════
#  Build the sheet
# ══════════════════════════════════════════════════════════════
row = 1

# ── Title row (merged across B:C) ──
ws.merge_cells('B1:C1')
cell_t = ws['B1']
cell_t.value = "TABLE I"
cell_t.font  = ft_title
cell_t.alignment = al_center
ws.row_dimensions[row].height = 28
row += 1

# ── Subtitle row (merged, with salmon fill) ──
ws.merge_cells('B2:C2')
cell_s = ws['B2']
cell_s.value = "NOTATIONS AND EXPLANATIONS"
cell_s.font  = Font(name='Times New Roman', size=11, bold=True, color='000000')
cell_s.alignment = al_center
cell_s.fill = fill_header
ws.row_dimensions[row].height = 24
# Apply fill to both cells in the merge range
ws['C2'].fill = fill_header
row += 1

# ── Thick rule below subtitle ──
for col in ('B', 'C'):
    ws[f'{col}{row-1}'].border = Border(bottom=thick)

# ── Empty separator row ──
ws.row_dimensions[row].height = 6
row += 1

# ── Column headers ──
ws[f'B{row}'].value = "Notation"
ws[f'B{row}'].font  = ft_header
ws[f'B{row}'].alignment = al_center
ws[f'B{row}'].border = Border(top=thick, bottom=medium)

ws[f'C{row}'].value = "Explanation"
ws[f'C{row}'].font  = ft_header
ws[f'C{row}'].alignment = al_center
ws[f'C{row}'].border = Border(top=thick, bottom=medium)
ws.row_dimensions[row].height = 26
row += 1

# ── Data rows ──
for i, (notation, explanation) in enumerate(notations):
    r = row + i
    ws[f'B{r}'].value = notation
    ws[f'B{r}'].font  = ft_body   # italic for math symbols
    ws[f'B{r}'].alignment = al_center

    ws[f'C{r}'].value = explanation
    ws[f'C{r}'].font  = ft_body_r  # normal for descriptions
    ws[f'C{r}'].alignment = al_left

    ws.row_dimensions[r].height = 22

last_data_row = row + len(notations) - 1

# ── Thick bottom rule ──
for col in ('B', 'C'):
    ws[f'{col}{last_data_row}'].border = Border(bottom=thick)

# ── Print settings ──
ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(
    fitToPage=True
)
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 0

out_path = '/home/user/sssdtcn/Figures4paper/notation_table.xlsx'
wb.save(out_path)
print(f'Saved to {out_path}')
