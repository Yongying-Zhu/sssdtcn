"""
评估所有插补方法并生成Excel表格
"""
import numpy as np
import pickle
import torch
import sys
import warnings
warnings.filterwarnings('ignore')

sys.path.append('/home/zhu/sssdtcn/compare')

def calc_mae(pred, true, mask):
    diff = np.abs(pred - true)
    return np.sum(diff * mask) / np.sum(mask)

def calc_rmse(pred, true, mask):
    diff = (pred - true) ** 2
    return np.sqrt(np.sum(diff * mask) / np.sum(mask))

# ============================================
# 加载模型
# ============================================
def load_median_model(dataset_name):
    path = f'/home/zhu/sssdtcn/compare/models/median_{dataset_name}_model.pkl'
    with open(path, 'rb') as f:
        return pickle.load(f)['medians']

def load_last_model(dataset_name):
    path = f'/home/zhu/sssdtcn/compare/models/last_{dataset_name}_model.pkl'
    with open(path, 'rb') as f:
        return pickle.load(f)['global_means']

def load_mrnn_model(dataset_name, n_features):
    from train_mrnn_custom import MRNN
    path = f'/home/zhu/sssdtcn/compare/models/mrnn_{dataset_name}_model.pth'
    checkpoint = torch.load(path, map_location='cuda', weights_only=False)
    model = MRNN(n_features, hidden_size=128, dropout=0.15)
    model.load_state_dict(checkpoint['model_state_dict'])
    model.eval()
    model.cuda()
    return model

def load_gpvae_model(dataset_name, n_features):
    from train_gpvae_custom import GPVAE
    path = f'/home/zhu/sssdtcn/compare/models/gpvae_{dataset_name}_model.pth'
    checkpoint = torch.load(path, map_location='cuda', weights_only=False)
    model = GPVAE(n_features, seq_len=60, latent_dim=32, hidden_dim=128)
    model.load_state_dict(checkpoint['model_state_dict'])
    model.eval()
    model.cuda()
    return model

def load_pypots_model(method_name, dataset_name, n_features):
    """创建PyPOTS模型实例并加载权重"""
    from pypots.imputation import SAITS, BRITS, Transformer
    from pypots.optim import Adam
    
    path = f'/home/zhu/sssdtcn/compare/models/{method_name}_{dataset_name}_model.pypots'
    
    # 创建模型实例（参数需与训练时一致）
    optimizer = Adam(lr=1e-4)
    
    if method_name == 'saits':
        model = SAITS(
            n_steps=60, n_features=n_features,
            n_layers=2, d_model=128, n_heads=4,
            d_k=32, d_v=32, d_ffn=256, dropout=0.1,
            epochs=1, batch_size=32, optimizer=optimizer, device='cuda'
        )
    elif method_name == 'brits':
        model = BRITS(
            n_steps=60, n_features=n_features,
            rnn_hidden_size=128,
            epochs=1, batch_size=32, optimizer=optimizer, device='cuda'
        )
    elif method_name == 'transformer':
        model = Transformer(
            n_steps=60, n_features=n_features,
            n_layers=2, d_model=128, n_heads=4,
            d_k=32, d_v=32, d_ffn=256, dropout=0.1,
            epochs=1, batch_size=32, optimizer=optimizer, device='cuda'
        )
    
    # 加载权重
    model.load(path)
    return model

# ============================================
# 插补函数
# ============================================
def impute_median(data, mask, medians):
    result = data.copy()
    for f in range(data.shape[-1]):
        result[:, :, f] = np.where(mask[:, :, f], data[:, :, f], medians[f])
    return result

def impute_last(data, mask, global_means):
    result = data.copy()
    batch, seq_len, features = data.shape
    for b in range(batch):
        for f in range(features):
            last_valid = global_means[f]
            for t in range(seq_len):
                if mask[b, t, f]:
                    last_valid = data[b, t, f]
                result[b, t, f] = last_valid if not mask[b, t, f] else data[b, t, f]
    return result

def impute_mrnn(data, mask, model):
    device = next(model.parameters()).device
    x = torch.FloatTensor(data).to(device)
    m = torch.BoolTensor(mask).to(device)
    masked_x = x * m.float()
    with torch.no_grad():
        output = model(masked_x, m)
    result = torch.where(m, x, output)
    return result.cpu().numpy()

def impute_gpvae(data, mask, model):
    device = next(model.parameters()).device
    x = torch.FloatTensor(data).to(device)
    m = torch.BoolTensor(mask).to(device)
    masked_x = x * m.float()
    with torch.no_grad():
        recon, _, _ = model(masked_x, m)
    result = torch.where(m, x, recon)
    return result.cpu().numpy()

def impute_pypots(data, mask, model):
    """使用PyPOTS模型进行插补"""
    X = data.copy()
    X[~mask] = np.nan
    result = model.impute({'X': X})
    return result

# ============================================
# 评估
# ============================================
def evaluate_method(method_name, test_data, missing_ratio, model_or_params):
    np.random.seed(42)
    mask = np.random.rand(*test_data.shape) > missing_ratio
    missing_mask = ~mask
    
    if method_name == 'median':
        imputed = impute_median(test_data, mask, model_or_params)
    elif method_name == 'last':
        imputed = impute_last(test_data, mask, model_or_params)
    elif method_name == 'mrnn':
        imputed = impute_mrnn(test_data, mask, model_or_params)
    elif method_name == 'gpvae':
        imputed = impute_gpvae(test_data, mask, model_or_params)
    elif method_name in ['saits', 'brits', 'transformer']:
        imputed = impute_pypots(test_data, mask, model_or_params)
    
    mae = calc_mae(imputed, test_data, missing_mask)
    rmse = calc_rmse(imputed, test_data, missing_mask)
    return mae, rmse

def evaluate_all_methods(dataset_name, test_data, n_features):
    missing_ratios = [0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8]
    methods = ['Median', 'Last', 'M-RNN', 'GP-VAE', 'BRITS', 'Transformer', 'SAITS']
    results = {m: {'MAE': [], 'RMSE': []} for m in methods}
    
    print(f"Loading models for {dataset_name}...")
    models = {
        'Median': ('median', load_median_model(dataset_name)),
        'Last': ('last', load_last_model(dataset_name)),
        'M-RNN': ('mrnn', load_mrnn_model(dataset_name, n_features)),
        'GP-VAE': ('gpvae', load_gpvae_model(dataset_name, n_features)),
        'SAITS': ('saits', load_pypots_model('saits', dataset_name, n_features)),
        'BRITS': ('brits', load_pypots_model('brits', dataset_name, n_features)),
        'Transformer': ('transformer', load_pypots_model('transformer', dataset_name, n_features)),
    }
    print("All models loaded.")
    
    for ratio in missing_ratios:
        print(f"\nEvaluating at {int(ratio*100)}% missing...")
        for method in methods:
            key, model = models[method]
            mae, rmse = evaluate_method(key, test_data, ratio, model)
            results[method]['MAE'].append(mae)
            results[method]['RMSE'].append(rmse)
            print(f"  {method}: MAE={mae:.4f}, RMSE={rmse:.4f}")
    
    return results, missing_ratios

# ============================================
# Excel生成
# ============================================
def create_excel(sru_res, deb_res, sru_ours, deb_ours, ratios, path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side
    
    wb = Workbook()
    methods = ['Median', 'Last', 'M-RNN', 'GP-VAE', 'BRITS', 'Transformer', 'SAITS', 'Ours']
    hfont = Font(bold=True)
    bfont = Font(bold=True)
    center = Alignment(horizontal='center', vertical='center')
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                   top=Side(style='thin'), bottom=Side(style='thin'))
    
    def fill_sheet(ws, name, res, ours):
        ws['A1'] = 'Dataset'; ws['B1'] = 'Model'
        ws['A1'].font = ws['B1'].font = hfont
        ws['A1'].alignment = ws['B1'].alignment = center
        
        col = 3
        for r in ratios:
            ws.cell(1, col, f"{int(r*100)}% missing").font = hfont
            ws.cell(1, col).alignment = center
            ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col+1)
            ws.cell(2, col, 'MAE').font = hfont; ws.cell(2, col).alignment = center
            ws.cell(2, col+1, 'RMSE').font = hfont; ws.cell(2, col+1).alignment = center
            col += 2
        
        ws.merge_cells(start_row=3, start_column=1, end_row=3+len(methods)-1, end_column=1)
        ws.cell(3, 1, name.upper()).alignment = center
        
        for i, m in enumerate(methods):
            row = 3 + i
            ws.cell(row, 2, m).alignment = center
            if m == 'Ours': ws.cell(row, 2).font = bfont
            col = 3
            for j in range(len(ratios)):
                if m == 'Ours':
                    mae, rmse = ours['MAE'][j], ours['RMSE'][j]
                    ws.cell(row, col).font = ws.cell(row, col+1).font = bfont
                else:
                    mae, rmse = res[m]['MAE'][j], res[m]['RMSE'][j]
                ws.cell(row, col, round(mae, 3)).alignment = center
                ws.cell(row, col+1, round(rmse, 3)).alignment = center
                col += 2
        
        for r in range(1, 3+len(methods)):
            for c in range(1, 3+len(ratios)*2):
                ws.cell(r, c).border = border
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 12
    
    ws1 = wb.active; ws1.title = 'SRU'
    fill_sheet(ws1, 'SRU', sru_res, sru_ours)
    
    ws2 = wb.create_sheet('Debutanizer')
    fill_sheet(ws2, 'Debutanizer', deb_res, deb_ours)
    
    wb.save(path)
    print(f"\nExcel saved: {path}")

def main():
    print("="*60)
    print("评估所有插补方法")
    print("="*60)
    
    sru_ours = {
        'MAE': [0.077449, 0.071798, 0.070129, 0.072755, 0.081729, 0.103772, 0.156999],
        'RMSE': [0.102070, 0.096791, 0.096379, 0.101043, 0.115070, 0.148243, 0.228235]
    }
    deb_ours = {
        'MAE': [0.275630, 0.282879, 0.290400, 0.300365, 0.308491, 0.319918, 0.342980],
        'RMSE': [0.378416, 0.388883, 0.399530, 0.414309, 0.428030, 0.447579, 0.485922]
    }
    
    sru_test = np.load('/home/zhu/sssdtcn/compare/models/sru_test_data.npy')
    deb_test = np.load('/home/zhu/sssdtcn/compare/models/debutanizer_test_data.npy')
    print(f"SRU test: {sru_test.shape}, Debutanizer test: {deb_test.shape}")
    
    print("\n" + "="*60)
    print("Evaluating SRU")
    print("="*60)
    sru_res, ratios = evaluate_all_methods('sru', sru_test, sru_test.shape[-1])
    
    print("\n" + "="*60)
    print("Evaluating Debutanizer")
    print("="*60)
    deb_res, _ = evaluate_all_methods('debutanizer', deb_test, deb_test.shape[-1])
    
    create_excel(sru_res, deb_res, sru_ours, deb_ours, ratios,
                '/home/zhu/sssdtcn/compare/imputation_results.xlsx')
    
    print("\n" + "="*60)
    print("完成!")
    print("="*60)

if __name__ == '__main__':
    main()
